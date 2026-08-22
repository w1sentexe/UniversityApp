"""Разбор расписания занятий из выгрузки Excel в структуру по группам.

Формат исходника (проверен на файле УИТС, все листы размечены одинаково):

* лист = курс, шапка с названиями групп — строка со словом «Дни» в столбце A;
* столбец A — день недели, объединён на весь свой блок строк;
* столбец B — пара, объединена ровно на ДВЕ строки: верхняя — числитель,
  нижняя — знаменатель;
* столбцы с третьего — группы. Группа занимает одну колонку (не делится) либо
  две объединённые (левая — первая подгруппа, правая — вторая);
* лекция на поток — одна ячейка, объединённая по колонкам нескольких групп;
  её значение лежит в левой верхней ячейке, остальные читаются как None,
  поэтому объединения приходится разворачивать вручную.

Ячейка занятия выглядит так:

    пр.Иностранный язык
    Мирошниченко Е.Н.   21           Павлова С.В.   11и

Первая строка — тип и название, вторая — преподаватель и аудитория. Если во
второй строке две пары «преподаватель + аудитория», значит занятие всё же
разведено по подгруппам внутри одной объединённой ячейки.

Модуль ничего не знает ни про БД, ни про сеть: на вход путь к файлу, на выход
словарь {группа: расписание}.
"""

import re
from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.logging_config import get_logger

log = get_logger(__name__)

# Столбец A шапки; по нему находим строку с названиями групп.
_HEADER_MARKER = "Дни"
# Данные начинаются с третьей колонки: первые две заняты днём и временем.
_FIRST_GROUP_COL = 3

# Префикс в начале ячейки → тип занятия.
_LESSON_TYPES = {
    "л": "lecture",
    "пр": "practice",
    "лаб": "lab",
    "сем": "seminar",
}

_DAY_KEYS = {
    "ПОНЕДЕЛЬНИК": "monday",
    "ВТОРНИК": "tuesday",
    "СРЕДА": "wednesday",
    "ЧЕТВЕРГ": "thursday",
    "ПЯТНИЦА": "friday",
    "СУББОТА": "saturday",
}

_WEEKS = ("numerator", "denominator")

# «Мирошниченко Е.Н.   21» — фамилия с инициалами, затем через отступ аудитория.
# Аудитория необязательна и может быть с буквой: 309б, 11и, 020.
_TEACHER_ROOM = re.compile(r"([А-ЯЁA-Z][^\s]*(?:\s+[А-ЯЁA-Z]\.){0,2}[^\s]*)\s{2,}(\S+)")
# Разделитель двух пар «преподаватель + аудитория» внутри одной ячейки.
_PAIR_SPLIT = re.compile(r"\s{4,}")


class ScheduleParseError(RuntimeError):
    """Файл не удалось разобрать: разметка не совпала с ожидаемой."""


def _merged_lookup(sheet: Worksheet) -> dict[tuple[int, int], tuple[int, int, int]]:
    """Карта «ячейка → (строка якоря, колонка якоря, ширина объединения)».

    Нужна, чтобы развернуть объединения: значение лежит только в левой верхней
    ячейке, а занятие относится ко всем накрытым колонкам.
    """
    lookup: dict[tuple[int, int], tuple[int, int, int]] = {}
    for rng in sheet.merged_cells.ranges:
        width = rng.max_col - rng.min_col + 1
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                lookup[(row, col)] = (rng.min_row, rng.min_col, width)
    return lookup


def _cell_value(sheet: Worksheet, merged: dict, row: int, col: int) -> tuple[Any, int]:
    """Значение ячейки с учётом объединений и ширина накрывающего блока."""
    anchor = merged.get((row, col))
    if anchor is None:
        return sheet.cell(row, col).value, 1
    anchor_row, anchor_col, width = anchor
    return sheet.cell(anchor_row, anchor_col).value, width


def _find_header_row(sheet: Worksheet) -> int:
    for row in range(1, 20):
        if str(sheet.cell(row, 1).value or "").strip() == _HEADER_MARKER:
            return row
    raise ScheduleParseError(f"На листе {sheet.title!r} не найдена строка шапки со словом {_HEADER_MARKER!r}")


def _read_groups(sheet: Worksheet, header_row: int, merged: dict) -> list[tuple[str, int, int]]:
    """Список (название группы, первая колонка, число колонок).

    Две колонки означают деление на подгруппы.
    """
    groups: list[tuple[str, int, int]] = []
    col = _FIRST_GROUP_COL
    while col <= sheet.max_column:
        value, width = _cell_value(sheet, merged, header_row, col)
        name = str(value).strip() if value is not None else ""
        if name:
            groups.append((name, col, width))
        col += width
    return groups


def _read_day_blocks(sheet: Worksheet) -> list[tuple[str, int, int]]:
    """Дни недели: (ключ дня, первая строка блока, последняя строка)."""
    blocks: list[tuple[str, int, int]] = []
    for rng in sorted(sheet.merged_cells.ranges, key=lambda r: r.min_row):
        if rng.min_col != 1 or rng.max_row == rng.min_row:
            continue
        title = str(sheet.cell(rng.min_row, 1).value or "").strip().upper()
        key = _DAY_KEYS.get(title)
        if key:
            blocks.append((key, rng.min_row, rng.max_row))
    return blocks


def _read_slots(sheet: Worksheet, first_row: int, last_row: int) -> Iterator[tuple[str, str, int]]:
    """Пары внутри дня: (начало «08:00», конец «09:35», строка числителя).

    Строка знаменателя — следующая за ней: пара всегда занимает две строки.
    """
    for row in range(first_row, last_row + 1):
        raw = sheet.cell(row, 2).value
        if raw is None or not str(raw).strip():
            continue
        parts = [p.strip().replace(".", ":") for p in str(raw).strip().split("-")]
        start = parts[0]
        end = parts[1] if len(parts) > 1 else ""
        yield start, end, row


def _split_teacher_room(text: str) -> list[tuple[str, str | None]]:
    """Разбирает строку с преподавателями и аудиториями.

    Возвращает список пар: одна — занятие общее, две — разведено по подгруппам
    прямо внутри объединённой ячейки.
    """
    chunks = [c.strip() for c in _PAIR_SPLIT.split(text.strip()) if c.strip()]
    pairs: list[tuple[str, str | None]] = []
    for chunk in chunks:
        match = _TEACHER_ROOM.match(chunk)
        if match:
            pairs.append((match.group(1).strip(), match.group(2).strip()))
        else:
            pairs.append((chunk, None))
    return pairs or [(text.strip(), None)]


def _parse_cell(raw: str) -> list[dict[str, Any]] | None:
    """Ячейка → одно занятие или два (когда подгруппы разведены внутри неё)."""
    lines = [line.strip() for line in str(raw).split("\n") if line.strip()]
    if not lines:
        return None

    head = lines[0]
    match = re.match(r"([А-Яа-яA-Za-z]+)\.(.+)", head)
    if match and match.group(1).lower() in _LESSON_TYPES:
        lesson_type = _LESSON_TYPES[match.group(1).lower()]
        name = match.group(2).strip()
    else:
        # Без префикса приходит, например, «Общая физическая подготовка».
        lesson_type = "other"
        name = head

    if len(lines) < 2:
        return [{"type": lesson_type, "name": name, "teacherName": None, "classroom": None}]

    return [
        {"type": lesson_type, "name": name, "teacherName": teacher, "classroom": room}
        for teacher, room in _split_teacher_room(lines[1])
    ]


def _empty_schedule() -> dict[str, dict[str, dict[str, dict]]]:
    return {week: {day: {"schedule": {}} for day in _DAY_KEYS.values()} for week in _WEEKS}


def parse_schedule(path: str | Path) -> dict[str, dict]:
    """Разбирает файл расписания и возвращает {название группы: расписание}.

    Расписание группы имеет вид
    {"numerator": {"monday": {"schedule": {"08:00": [занятие, ...]}}}, "denominator": {...}},
    где ключ — начало пары, а конец лежит в самом занятии полем endTime.

    Значение по времени — список: в одной паре у подгрупп бывают разные занятия,
    и в один объект они не помещаются.
    """
    import openpyxl  # локальный импорт: нужен только скрипту импорта, не рантайму API

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=False)
    result: dict[str, dict] = {}
    lessons_total = 0

    for sheet in workbook.worksheets:
        try:
            header_row = _find_header_row(sheet)
        except ScheduleParseError as exc:
            log.warning("Sheet skipped", sheet=sheet.title, reason=str(exc))
            continue

        merged = _merged_lookup(sheet)
        groups = _read_groups(sheet, header_row, merged)
        days = _read_day_blocks(sheet)
        log.debug("Sheet layout", sheet=sheet.title, groups=len(groups), days=len(days))

        for name, first_col, width in groups:
            schedule = result.setdefault(name, _empty_schedule())

            for day_key, day_first, day_last in days:
                for start, end, slot_row in _read_slots(sheet, day_first, day_last):
                    for offset, week in enumerate(_WEEKS):
                        row = slot_row + offset
                        if row > day_last:
                            continue
                        lessons = _lessons_at(sheet, merged, row, first_col, width, name, end)
                        if lessons:
                            schedule[week][day_key]["schedule"][start] = lessons
                            lessons_total += len(lessons)

    log.info("Schedule parsed", groups=len(result), lessons=lessons_total)
    return result


def _lessons_at(
    sheet: Worksheet,
    merged: dict,
    row: int,
    first_col: int,
    width: int,
    group_name: str,
    end_time: str,
) -> list[dict[str, Any]]:
    """Занятия одной группы в конкретной строке (числитель или знаменатель)."""
    # Собираем значения по колонкам группы, схлопывая одинаковые: объединённая
    # на обе подгруппы ячейка даёт одно и то же значение в каждой колонке.
    seen: list[tuple[Any, int]] = []
    for col in range(first_col, first_col + width):
        value, _ = _cell_value(sheet, merged, row, col)
        if value is None or not str(value).strip():
            continue
        if not seen or seen[-1][0] != value:
            seen.append((value, col))

    lessons: list[dict[str, Any]] = []
    for value, col in seen:
        parsed = _parse_cell(value)
        if not parsed:
            continue
        for index, lesson in enumerate(parsed):
            # Подгруппа определяется либо позицией колонки внутри группы,
            # либо порядком пары «преподаватель + аудитория» в общей ячейке.
            if len(parsed) > 1:
                subgroup = [index + 1]
            elif len(seen) > 1:
                subgroup = [col - first_col + 1]
            else:
                subgroup = [1, 2]
            # endTime дублируется в каждом занятии слота намеренно: так документ
            # самодостаточен и клиенту не нужна зашитая таблица звонков.
            lessons.append({**lesson, "endTime": end_time, "group": group_name, "subgroup": subgroup})
    return lessons
